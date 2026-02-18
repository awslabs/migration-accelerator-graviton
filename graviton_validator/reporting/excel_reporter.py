"""
Excel report generator for Graviton compatibility analysis results.
"""

from typing import Dict, List, Optional, Any
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Create dummy classes for type hints when openpyxl is not available
    class Workbook:
        pass

from .base import ReportGenerator
from .json_reporter import JSONReporter
from ..models import AnalysisResult


class ExcelReporter(ReportGenerator):
    """
    Excel report generator that creates comprehensive Excel workbooks.
    Uses JSONReporter internally for data structuring.
    """
    
    def __init__(self, include_charts: bool = True):
        """
        Initialize Excel reporter.
        
        Args:
            include_charts: Whether to include charts and graphs
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel reporting. "
                "Install it with: pip install openpyxl"
            )
        
        self.include_charts = include_charts
        self.json_reporter = JSONReporter(include_metadata=True)
        
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.compatible_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.incompatible_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.unknown_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    def generate_report(self, analysis_result: AnalysisResult, output_path: Optional[str] = None) -> str:
        """
        Generate Excel report from analysis results.
        
        Args:
            analysis_result: AnalysisResult to generate report from
            output_path: Optional path to write report to file
            
        Returns:
            Path to generated Excel file or binary content info
        """
        # Get structured data from JSON reporter
        data = self.json_reporter.get_structured_data(analysis_result)
        
        # Create workbook
        workbook = self._create_workbook(data, analysis_result)
        
        # Save to file or return binary content
        if output_path:
            workbook.save(output_path)
            return f"Excel report saved to {output_path}"
        else:
            # Return workbook as bytes for testing
            buffer = io.BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            return f"Excel workbook generated ({len(buffer.getvalue())} bytes)"
    
    def get_format_name(self) -> str:
        """Get the name of the report format."""
        return "excel"
    
    def _create_workbook(self, data: Dict[str, Any], analysis_result: AnalysisResult = None) -> Workbook:
        """
        Create complete Excel workbook from structured data.
        
        Args:
            data: Structured report data from JSON reporter
            
        Returns:
            Configured Excel workbook
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb, data)
        self._create_unique_components_sheet(wb, data, analysis_result)
        self._create_detailed_results_sheet(wb, data, analysis_result)
        self._create_recommendations_sheet(wb, data)
        
        if data["errors"]:
            self._create_errors_sheet(wb, data)
        
        # Set active sheet to summary
        wb.active = wb["Summary"]
        
        return wb
    
    def _create_summary_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create summary overview sheet."""
        ws = workbook.create_sheet("Summary")
        
        # Title
        ws["A1"] = "Graviton Compatibility Analysis Summary"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:F1")
        
        # Metadata section
        if "metadata" in data:
            metadata = data["metadata"]
            ws["A3"] = "Report Information"
            ws["A3"].font = Font(bold=True)
            
            current_row = 4
            ws[f"A{current_row}"] = "Generated:"
            ws[f"B{current_row}"] = metadata["generated_at"]
            current_row += 1
            
            ws[f"A{current_row}"] = "Generator:"
            ws[f"B{current_row}"] = f"{metadata['generator']} v{metadata['version']}"
            current_row += 1
            
            if metadata.get("detected_os"):
                ws[f"A{current_row}"] = "Detected OS:"
                ws[f"B{current_row}"] = metadata["detected_os"]
                current_row += 1
            
            if metadata.get("sbom_file"):
                ws[f"A{current_row}"] = "SBOM File:"
                ws[f"B{current_row}"] = metadata["sbom_file"]
                current_row += 1
            
            ws[f"A{current_row}"] = "Processing Time:"
            ws[f"B{current_row}"] = f"{data['summary']['processing_time_seconds']}s"
        
        # OS Summary (if available)
        summary = data["summary"]
        current_row = 8
        
        if "os_summary" in summary:
            os_summary = summary["os_summary"]
            ws[f"A{current_row}"] = "Operating System Analysis"
            ws[f"A{current_row}"].font = Font(bold=True)
            current_row += 1
            
            # OS details
            ws[f"A{current_row}"] = "Detected OS:"
            ws[f"B{current_row}"] = os_summary["detected_os"]
            current_row += 1
            
            ws[f"A{current_row}"] = "System Packages:"
            ws[f"B{current_row}"] = os_summary["system_packages"]
            current_row += 1
            
            ws[f"A{current_row}"] = "Application Packages:"
            ws[f"B{current_row}"] = os_summary["application_packages"]
            current_row += 1
            
            ws[f"A{current_row}"] = "OS Graviton Compatible:"
            ws[f"B{current_row}"] = "Yes" if os_summary["os_compatible"] else "No"
            if os_summary["os_compatible"]:
                ws[f"B{current_row}"].fill = self.compatible_fill
            else:
                ws[f"B{current_row}"].fill = self.incompatible_fill
            current_row += 1
            
            ws[f"A{current_row}"] = "System Package %:"
            ws[f"B{current_row}"] = f"{os_summary['system_package_percentage']}%"
            current_row += 2
        
        # Summary metrics
        ws[f"A{current_row}"] = "Compatibility Summary"
        ws[f"A{current_row}"].font = Font(bold=True)
        current_row += 1
        
        # Create summary table
        headers = ["Metric", "Count", "Percentage", "Recommended Action"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        current_row += 1
        
        # Summary data
        total = summary["total_components"]
        needs_verification = summary.get("needs_verification", 0)
        needs_upgrade = summary.get("needs_upgrade", 0)
        needs_version_verification = summary.get("needs_version_verification", 0)
        summary_data = [
            ("Total Components", total, "100%", "Review individual components below"),
            ("Compatible", summary["compatible"], f"{summary['compatibility_rate']}%", "Package works on Graviton without changes. Safe to migrate."),
            ("Incompatible", summary["incompatible"], f"{round(summary['incompatible']/total*100, 1) if total > 0 else 0}%", "Package has known issues on Graviton. Find alternatives or wait for updates."),
            ("Needs Upgrade", needs_upgrade, f"{round(needs_upgrade/total*100, 1) if total > 0 else 0}%", "Current version doesn't support Graviton, but newer versions do. Upgrade recommended."),
            ("Needs Verification", needs_verification, f"{round(needs_verification/total*100, 1) if total > 0 else 0}%", "Package likely works on Graviton but requires testing to confirm compatibility."),
            ("Needs Version Verification", needs_version_verification, f"{round(needs_version_verification/total*100, 1) if total > 0 else 0}%", "Software is Graviton-compatible but version information is missing. Verify your version meets requirements."),
            ("Unknown", summary["unknown"], f"{round(summary['unknown']/total*100, 1) if total > 0 else 0}%", "No compatibility information available. Manual testing required before migration.")
        ]
        
        for i, (metric, count, percentage, action) in enumerate(summary_data):
            row = current_row + i
            ws.cell(row=row, column=1, value=metric).border = self.border
            ws.cell(row=row, column=2, value=count).border = self.border
            ws.cell(row=row, column=3, value=percentage).border = self.border
            ws.cell(row=row, column=4, value=action).border = self.border
            
            # Apply conditional formatting
            if metric == "Compatible":
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = self.compatible_fill
            elif metric == "Incompatible":
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = self.incompatible_fill
            elif metric == "Needs Upgrade":
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # Light orange
            elif metric == "Needs Verification":
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = self.unknown_fill
            elif metric == "Needs Version Verification":
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # Light lavender
            elif metric == "Unknown":
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = self.unknown_fill
        
        current_row += len(summary_data) + 2
        
        # SBOM breakdown
        ws[f"A{current_row}"] = "Analysis by SBOM File"
        ws[f"A{current_row}"].font = Font(bold=True)
        current_row += 1
        
        sbom_headers = ["SBOM File", "Detected OS", "OS Support", "Compatible", "Incompatible", "Needs Upgrade", "Needs Verification", "Needs Version Verification", "Unknown", "Total"]
        for col, header in enumerate(sbom_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        current_row += 1
        
        sbom_breakdown = data["statistics"]["sbom_breakdown"]
        for i, (sbom_file, counts) in enumerate(sbom_breakdown.items()):
            row = current_row + i
            total_sbom = counts["compatible"] + counts["incompatible"] + counts.get("needs_upgrade", 0) + counts.get("needs_verification", 0) + counts.get("needs_version_verification", 0) + counts["unknown"]
            detected_os = counts.get("detected_os", "N/A")
            os_support = counts.get("os_support_status", "Unknown")
            sbom_data = [sbom_file, detected_os, os_support, counts["compatible"], counts["incompatible"], counts.get("needs_upgrade", 0), counts.get("needs_verification", 0), counts.get("needs_version_verification", 0), counts["unknown"], total_sbom]
            
            for col, value in enumerate(sbom_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col > 3:  # Numeric columns (now starting from column 4)
                    cell.alignment = self.center_alignment
                elif col == 3:  # OS Support column - add color coding
                    cell.alignment = self.center_alignment
                    if value == "Supported":
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif value == "Not Supported":
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Add charts if enabled
        if self.include_charts:
            self._add_summary_charts(ws, summary, len(sbom_breakdown))
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col_num in range(1, 7):  # Adjust first 6 columns
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            # Check all rows for this column
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    if hasattr(cell, 'value') and cell.value is not None and not hasattr(cell, 'coordinate'):
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except (TypeError, AttributeError):
                            # Skip cells with invalid values
                            pass
            
            adjusted_width = min(max(max_length + 2, 15), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set column D width to 13 for better layout
        ws.column_dimensions['D'].width = 13
    
    def _create_detailed_results_sheet(self, workbook: Workbook, data: Dict[str, Any], analysis_result: AnalysisResult = None):
        """Create detailed results sheet."""
        ws = workbook.create_sheet("Detailed Results")
        
        # Headers - conditionally include runtime analysis columns
        headers = [
            "Component Name",
            "Version",
            "Package Type",
            "SBOM File",
            "Graviton Support Status",
            # "Current Version Compatible",
            # "Min ARM64 Supported Version",
            # "Recommended Version for ARM64",
            # "Confidence",
            "Detected OS",
            "OS-Level Package",
            "Notes",
            "Matched Name"
        ]
        
        # Check if any components have runtime analysis data
        has_runtime_analysis = any(
            comp.get("properties", {}).get("runtime_analysis") == "true" 
            for comp in data["components"]
        )
        
        if has_runtime_analysis:
            headers.insert(-2, "Requested Version")
            headers.insert(-2, "Version Changed")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Component data
        components = data["components"]
        for row, component in enumerate(components, 2):
            compat = component["compatibility"]
            props = component.get("properties", {})
            
            # Ensure component name is never empty
            component_name = component.get("name") or "unknown"
            
            # Truncate SBOM file path to just filename
            from pathlib import Path
            sbom_file = Path(component.get("source_sbom", "unknown")).name
            
            # Clean up package type - strip version/environment info
            raw_type = component.get("type", "unknown")
            pkg_type = raw_type.split("-")[1] if raw_type.startswith("native-") else raw_type if raw_type in ('rpm', 'python', 'ruby', 'unknown', 'library', 'linux-kernel-module') else raw_type.split("-")[0]
            
            row_data = [
                component_name,
                component.get("version", "N/A"),
                pkg_type,
                sbom_file,
                compat["status"],
                # "Yes" if compat.get("current_version_supported") else "No",
                # compat.get("minimum_supported_version", "N/A"),
                # compat.get("recommended_version", "N/A"),
                # f"{int(compat.get('confidence_level', 0) * 100)}%" if compat.get("confidence_level") is not None else "N/A",
                props.get("detected_os") or props.get("sbom_detected_os") or (analysis_result.detected_os if analysis_result else None) or "N/A",
                "System" if props.get("os_system_package") == "true" else "Application",
                compat.get("notes", ""),
                component.get("matched_name", "")
            ]
            
            # Add runtime analysis columns if present
            if has_runtime_analysis:
                # Insert before notes and matched_name
                row_data.insert(-2, props.get("original_version", "N/A"))
                row_data.insert(-2, "Yes" if props.get("fallback_used") == "true" else "No")
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                
                # Apply status-based formatting
                if col == 5:  # Status column
                    if value == "compatible":
                        cell.fill = self.compatible_fill
                    elif value == "incompatible":
                        cell.fill = self.incompatible_fill
                    elif value == "needs_upgrade":
                        cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # Light orange
                    elif value == "needs_verification":
                        cell.fill = self.unknown_fill
                    elif value == "needs_version_verification":
                        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # Light lavender
                    elif value == "unknown":
                        cell.fill = self.unknown_fill
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        max_cols = len(headers)
        for col_num in range(1, max_cols + 1):
            max_length = 15  # Minimum width
            column_letter = get_column_letter(col_num)
            
            # Check all rows for this column
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if hasattr(cell, 'value') and cell.value is not None:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except (TypeError, AttributeError):
                        # Skip cells with invalid values
                        pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_unique_components_sheet(self, workbook: Workbook, data: Dict[str, Any], analysis_result: AnalysisResult = None):
        """Create unique components sheet (deduplicated by name+version)."""
        ws = workbook.create_sheet("Unique Components")
        
        # Headers - same as detailed results but without SBOM File column
        headers = [
            "Component Name",
            "Version",
            "Package Type",
            "Graviton Support Status",
            "Detected OS",
            "OS-Level Package",
            "Notes",
            "Matched Name"
        ]
        
        # Check if any components have runtime analysis data
        has_runtime_analysis = any(
            comp.get("properties", {}).get("runtime_analysis") == "true" 
            for comp in data["components"]
        )
        
        if has_runtime_analysis:
            headers.insert(-2, "Requested Version")
            headers.insert(-2, "Version Changed")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Deduplicate components by name+version
        unique_components = {}
        for component in data["components"]:
            key = (component.get("name", "unknown"), component.get("version", "N/A"))
            if key not in unique_components:
                unique_components[key] = component
        
        # Component data
        for row, component in enumerate(unique_components.values(), 2):
            compat = component["compatibility"]
            props = component.get("properties", {})
            
            # Ensure component name is never empty
            component_name = component.get("name") or "unknown"
            
            # Clean up package type
            raw_type = component.get("type", "unknown")
            pkg_type = raw_type.split("-")[1] if raw_type.startswith("native-") else raw_type if raw_type in ('rpm', 'python', 'ruby', 'unknown', 'library', 'linux-kernel-module') else raw_type.split("-")[0]
            
            row_data = [
                component_name,
                component.get("version", "N/A"),
                pkg_type,
                compat["status"],
                props.get("detected_os") or props.get("sbom_detected_os") or (analysis_result.detected_os if analysis_result else None) or "N/A",
                "System" if props.get("os_system_package") == "true" else "Application",
                compat.get("notes", ""),
                component.get("matched_name", "")
            ]
            
            # Add runtime analysis columns if present
            if has_runtime_analysis:
                row_data.insert(-2, props.get("original_version", "N/A"))
                row_data.insert(-2, "Yes" if props.get("fallback_used") == "true" else "No")
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                
                # Apply status-based formatting
                if col == 4:  # Status column (shifted left without SBOM File)
                    if value == "compatible":
                        cell.fill = self.compatible_fill
                    elif value == "incompatible":
                        cell.fill = self.incompatible_fill
                    elif value == "needs_upgrade":
                        cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                    elif value == "needs_verification":
                        cell.fill = self.unknown_fill
                    elif value == "needs_version_verification":
                        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                    elif value == "unknown":
                        cell.fill = self.unknown_fill
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        max_cols = len(headers)
        for col_num in range(1, max_cols + 1):
            max_length = 15
            column_letter = get_column_letter(col_num)
            
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if hasattr(cell, 'value') and cell.value is not None:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except (TypeError, AttributeError):
                        pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_recommendations_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create recommendations sheet."""
        ws = workbook.create_sheet("Recommendations")
        
        # Title
        ws["A1"] = "Graviton Compatibility Recommendations"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:D1")
        
        row = 3
        
        # Deduplicate components by name+version
        components = data["components"]
        unique_components = {}
        for component in components:
            key = (component.get("name", "unknown"), component.get("version", "N/A"))
            if key not in unique_components:
                unique_components[key] = component
        
        unique_list = list(unique_components.values())
        
        # Components requiring upgrades
        components_with_upgrades = [
            c for c in unique_list 
            if c["compatibility"]["status"] in ["incompatible", "needs_upgrade"] 
            and c["compatibility"].get("recommended_version")
        ]
        
        if components_with_upgrades:
            ws[f"A{row}"] = "Components with Available Upgrades"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            
            # Headers
            upgrade_headers = ["Component", "Current Version", "Recommended Version", "Notes"]
            for col, header in enumerate(upgrade_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
            row += 1
            
            # Data
            for component in components_with_upgrades:
                upgrade_data = [
                    component["name"],
                    component.get("version", "N/A"),
                    component["compatibility"]["recommended_version"],
                    component["compatibility"].get("notes", "")
                ]
                
                for col, value in enumerate(upgrade_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.border
                row += 1
            
            row += 1
        
        # Components requiring investigation
        unknown_components = [c for c in unique_list if c["compatibility"]["status"] == "unknown"]
        incompatible_no_upgrade = [
            c for c in unique_list 
            if c["compatibility"]["status"] == "incompatible" 
            and not c["compatibility"].get("recommended_version")
        ]
        
        investigation_components = unknown_components + incompatible_no_upgrade
        
        if investigation_components:
            ws[f"A{row}"] = "Components Requiring Investigation"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            
            # Headers
            investigation_headers = ["Component", "Version", "Status", "Reason"]
            for col, header in enumerate(investigation_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
            row += 1
            
            # Data
            for component in investigation_components:
                status = component["compatibility"]["status"]
                reason = "Unknown compatibility" if status == "unknown" else "No upgrade path available"
                
                investigation_data = [
                    component["name"],
                    component.get("version", "N/A"),
                    status,
                    reason
                ]
                
                for col, value in enumerate(investigation_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.border
                    if col == 3:  # Status column
                        if value == "unknown":
                            cell.fill = self.unknown_fill
                        elif value == "incompatible":
                            cell.fill = self.incompatible_fill
                row += 1
            
            row += 2
        
        # Migration strategy
        ws[f"A{row}"] = "Recommended Migration Strategy"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        
        strategy_steps = [
            "1. Update components with known compatible versions",
            "2. Research and test unknown components in development environment",
            "3. Find alternatives for components without upgrade paths",
            "4. Validate entire application on Graviton instances",
            "5. Monitor performance and compatibility in production"
        ]
        
        for step in strategy_steps:
            ws[f"A{row}"] = step
            row += 1
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col_num in range(1, 5):  # Adjust columns A-D
            max_length = 15  # Minimum width
            column_letter = get_column_letter(col_num)
            
            # Check all rows for this column
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if hasattr(cell, 'value') and cell.value is not None:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except (TypeError, AttributeError):
                        # Skip cells with invalid values
                        pass
            
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_errors_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create errors and warnings sheet."""
        ws = workbook.create_sheet("Errors & Warnings")
        
        # Title
        ws["A1"] = "Processing Errors and Warnings"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:C1")
        
        # Headers
        headers = ["#", "Type", "Message"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Error data
        errors = data["errors"]
        for row, error in enumerate(errors, 4):
            error_type = "Warning" if error.lower().startswith("warning") else "Error"
            
            error_data = [row - 3, error_type, error]
            for col, value in enumerate(error_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                
                if col == 2 and error_type == "Error":
                    cell.fill = self.incompatible_fill
                elif col == 2 and error_type == "Warning":
                    cell.fill = self.unknown_fill
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col_num in range(1, 4):  # Adjust columns A-C
            max_length = 15  # Minimum width
            column_letter = get_column_letter(col_num)
            
            # Check all rows for this column
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if hasattr(cell, 'value') and cell.value is not None:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except (TypeError, AttributeError):
                        # Skip cells with invalid values
                        pass
            
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_summary_charts(self, worksheet, summary: Dict[str, Any], sbom_count: int):
        """Add charts to summary sheet."""
        try:
            # Pie chart for compatibility status
            pie_chart = PieChart()
            pie_chart.title = "Compatibility Status Distribution"
            
            # Set title to not overlay (appear above chart)
            if hasattr(pie_chart.title, 'overlay'):
                pie_chart.title.overlay = False
            
            # Data for pie chart - rows 18-23 (skip Total Components at row 17)
            # Compatible, Incompatible, Needs Upgrade, Needs Verification, Needs Version Verification, Unknown
            data = Reference(worksheet, min_col=2, min_row=18, max_row=23, max_col=2)
            labels = Reference(worksheet, min_col=1, min_row=18, max_row=23, max_col=1)
            
            pie_chart.add_data(data, titles_from_data=False)
            pie_chart.set_categories(labels)
            
            # Set chart size - larger to prevent legend overlap
            pie_chart.width = 20.81
            pie_chart.height = 12
            
            # Set legend to overlay (allows legend inside chart area)
            if hasattr(pie_chart, 'legend') and pie_chart.legend:
                pie_chart.legend.overlay = True
            
            # Position chart at column K, row 1
            worksheet.add_chart(pie_chart, "K1")
                
        except (ImportError, AttributeError, ValueError) as e:
            # Charts are optional, continue without them if there's an issue
            import logging
            logging.debug(f"Chart creation failed: {e}")